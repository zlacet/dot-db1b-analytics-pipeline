"""
generate_reports.py

Builds the two files in reports/ from the analytics tables:

    quarterly_report.xlsx   - 3 sheets:
                              1. Quarterly Trend - whole dataset (every
                                 route combined), one row per quarter,
                                 no domain grouping (it's not route-,
                                 price-, or carrier-specific)
                              2. Route Summary - every route, every
                                 column (passengers, average price, price
                                 per mile, passenger growth %, price
                                 change %, price-per-mile change %),
                                 sorted by passengers descending, no row
                                 limit.
                              3. Carrier Summary - every carrier, every
                                 column (total passengers, average price,
                                 average price per mile, average market
                                 share %), sorted by total passengers
                                 descending, no row limit.
                              There used to be several separate route
                              sheets (by volume / growth / price movement)
                              and several separate carrier sheets (by
                              passengers / price / share) - but route_summary
                              and carrier_summary already carry every metric
                              on every row, so those were the same rows
                              several times over, just resorted. One sheet
                              per entity now; the PDF still shows best/worst
                              by each individual metric (see below).
    executive_summary.pdf   - title, overview blurb, then the Quarterly
                              Trend table with its AI-written blurb right
                              underneath it. The remaining 12 sections
                              follow, flowing naturally (no forced page
                              break), grouped under three headings - Carrier
                              Performance, Price Trends, Route Performance,
                              each always starting on its own page. Price
                              Trends holds both the price-movement tables and
                              the price-efficiency tables. Growth/
                              price-movement tables are capped at
                              ai_layer/summarize.py's BEST_WORST_TOP_N
                              (currently 3); price-efficiency tables at
                              ROUTE_EFFICIENCY_TOP_N (currently 5); most
                              CARRIER tables at CARRIER_BEST_WORST_TOP_N
                              (currently 5), except Top Carriers by Market
                              Share at CARRIER_SHARE_TOP_N (currently 10) -
                              kept separate so each set of tables can be
                              resized independently. The "Top Routes by
                              Volume" table is the exception, capped at
                              --top instead.
                              Every table (all 12, plus Quarterly Trend) has
                              its own short blurb from ai_layer/summarize.py
                              placed directly underneath it - see
                              SECTION_ORDER there for the full section list.
                              A standalone "Overall Consensus" section, not
                              tied to any one table, always starts on its
                              own page at the end: a short opening overview,
                              then an Insight/Takeaway pair each for Carrier
                              Performance, Price Trends, and Route
                              Performance, then a single Quarterly Trends
                              paragraph - all describing overall patterns
                              only, no specific route/carrier names, since
                              those are already covered in the sections
                              above it.
                              These sections are derived in pandas (sort by
                              the relevant metric, then head()/tail(), plus
                              summarize.add_price_multiplier() for the
                              two Efficiency sections) from the single Route
                              Summary / Carrier Summary dataframes above,
                              plus one extra query for carrier price-per-mile
                              growth (no per-route/per-carrier metric query
                              beyond that).

Reuses the exact same query functions ai_layer/summarize.py's narrative
engines are built on (get_quarterly_trend, get_top_routes,
get_ranked_carriers, add_price_multiplier), so the report tables and
the narrative text can never disagree on numbers.

Reads directly from route_summary, carrier_summary, and
price_variance_summary — run this after pipeline/main.py has loaded the
current quarter's data.

Connection settings (same as the rest of the pipeline) can be overridden
with: PGHOST, PGPORT, PGDATABASE, PGUSER, PGPASSWORD

Usage:
    python generate_reports.py
    python generate_reports.py --top 15   # pdf's top-routes table shows top 15 instead of 10
"""

import argparse
import sys
from pathlib import Path

try:
    import pandas as pd
except ImportError:
    sys.exit("Missing dependency: pandas.\nInstall it with: pip install pandas --break-system-packages")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("Missing dependency: openpyxl.\nInstall it with: pip install openpyxl --break-system-packages")

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        KeepTogether,
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
except ImportError:
    sys.exit("Missing dependency: reportlab.\nInstall it with: pip install reportlab --break-system-packages")

REPORTS_DIR = Path(__file__).resolve().parent
XLSX_FILE = REPORTS_DIR / "quarterly_report.xlsx"
PDF_FILE = REPORTS_DIR / "executive_summary.pdf"

# Title printed at the top of executive_summary.pdf.
REPORT_TITLE = "Quarterly Airline Price & Demand Report"

# Static overview blurb printed right under the title, above the Quarterly
# Trends table - explains what the report is, not tied to any one quarter's
# data (unlike every other blurb in the report, which comes from
# ai_layer/summarize.py).
REPORT_OVERVIEW = (
    "This report summarizes U.S. domestic airline market activity for the quarter, "
    "covering passenger demand, average prices, price efficiency, and carrier competition "
    "across every route in the DOT's DB1B survey. It highlights the quarter's overall "
    "trend, the largest routes by volume, the fastest-growing and declining routes, the "
    "biggest price movements, which routes and carriers are priced efficiently relative to "
    "their peers, and how carriers stack up against each other, with analysis of each "
    "table written by Claude."
)

# Make ai_layer/summarize.py importable - reused for both the narrative
# text AND the ranking queries themselves (get_quarterly_trend,
# get_top_routes, get_ranked_routes, get_ranked_carriers,
# add_price_multiplier), so report tables and the narrative can never
# disagree on numbers.
sys.path.insert(0, str(REPORTS_DIR.parent / "ai_layer"))
import summarize  # noqa: E402

sys.path.insert(0, str(REPORTS_DIR.parent / "database"))
from db import get_connection  # noqa: E402

# Row cap for the PDF's Quarterly Trends table only - it grows by one row
# every quarter forever, so without a cap it eventually overruns the page.
# The xlsx's Quarterly Trend sheet stays unlimited, consistent with every
# other xlsx sheet (documented as "every row, no limit").
QUARTERLY_TREND_PDF_LIMIT = 15


# Which float-formatted columns each table shape needs.
QUARTERLY_TREND_FLOAT_COLS = [
    "total_passengers", "total_routes", "total_carriers", "average_price", "price_per_mile",
    "passenger_growth_pct", "price_change_pct",
]
ROUTE_FLOAT_COLS = [
    "passengers", "average_price", "price_per_mile",
    "passenger_growth_pct", "price_change_pct", "price_per_mile_change_pct",
]
# passenger_growth_pct is deliberately not in this list - the Price
# Efficiency tables drop that column entirely (see _pdf_sections) since it
# was pushing the table past the page's right margin and isn't part of the
# distance-adjusted efficiency story.
EFFICIENCY_FLOAT_COLS = ["passengers", "average_price", "price_per_mile", "distance", "price_multiplier"]
CARRIER_FLOAT_COLS = ["total_passengers", "avg_price", "avg_price_per_mile", "avg_route_share_pct"]
CARRIER_GROWTH_FLOAT_COLS = ["total_passengers", "avg_price_per_mile", "price_per_mile_growth_pct"]

# Which of the float-formatted columns above are dollar amounts,
# percentages, or multipliers, so _df_to_table can prefix/suffix them with
# $ / % / x instead of printing bare numbers. Every float column not listed
# in any of the three sets here (passenger/route/carrier counts, distance)
# is left as a plain comma-formatted number.
DOLLAR_COLS = {"average_price", "price_per_mile", "avg_price", "avg_price_per_mile"}
PERCENT_COLS = {
    "passenger_growth_pct", "price_change_pct", "price_per_mile_change_pct",
    "avg_route_share_pct", "price_per_mile_growth_pct",
}
# price_multiplier is a ratio versus distance-band peers, displayed with an
# "x" suffix (e.g. "2.75x") rather than as a $ amount or a percent.
MULTIPLIER_COLS = {"price_multiplier"}

# Domain groupings for the PDF: heading -> ordered list of section titles.
# "Quarterly Trend" isn't in here - it's printed on its own, before these.
# PRICE now holds both the route-shaped movement tables and the
# efficiency-shaped tables under one heading (no separate EFFICIENCY
# domain) - see SECTION_FLOAT_COLS below for how each table's column shape
# is looked up individually rather than per-domain, since PRICE's two table
# shapes differ.
DOMAIN_GROUPS = [
    (
        "Carrier Performance",
        [
            "Top Carriers by Passengers",
            "Bottom Carriers by Passengers",
            "Top Carriers by Price Efficiency",
            "Top Carriers by Price-Per-Mile Growth",
            "Top Carriers by Market Share",
        ],
    ),
    ("Price Trends", ["Best Price Efficiency", "Worst Price Efficiency", "Best Price Movement", "Worst Price Movement"]),
    ("Route Performance", ["Top Routes by Volume", "Best Passenger Growth", "Worst Passenger Growth"]),
]

# Which float-column set each individual table needs - looked up per
# section title (not per domain), since a domain (PRICE, in particular) can
# hold tables of more than one shape.
SECTION_FLOAT_COLS = {
    "Top Routes by Volume": ROUTE_FLOAT_COLS,
    "Best Passenger Growth": ROUTE_FLOAT_COLS,
    "Worst Passenger Growth": ROUTE_FLOAT_COLS,
    "Best Price Movement": ROUTE_FLOAT_COLS,
    "Worst Price Movement": ROUTE_FLOAT_COLS,
    "Best Price Efficiency": EFFICIENCY_FLOAT_COLS,
    "Worst Price Efficiency": EFFICIENCY_FLOAT_COLS,
    "Top Carriers by Passengers": CARRIER_FLOAT_COLS,
    "Bottom Carriers by Passengers": CARRIER_FLOAT_COLS,
    "Top Carriers by Price Efficiency": CARRIER_FLOAT_COLS,
    "Top Carriers by Price-Per-Mile Growth": CARRIER_GROWTH_FLOAT_COLS,
    "Top Carriers by Market Share": CARRIER_FLOAT_COLS,
}

# One-line business question printed under every table's sub-header
# (italic, smaller than the sub-header) - including Quarterly Trends, which
# isn't part of DOMAIN_GROUPS but still gets one. Answers "why does this
# table exist" for a first-time reader before they hit the numbers.
SECTION_QUESTIONS = {
    "Quarterly Trends": "How is the overall market trending this quarter?",
    "Top Routes by Volume": "Which routes carry the most passenger volume this quarter?",
    "Best Passenger Growth": "Which routes are gaining passengers fastest?",
    "Worst Passenger Growth": "Which routes are losing passengers fastest?",
    "Best Price Efficiency": "Which routes are priced highest relative to similar-distance routes?",
    "Worst Price Efficiency": "Which routes are priced lowest relative to similar-distance routes?",
    "Best Price Movement": "Which routes saw the largest price increases?",
    "Worst Price Movement": "Which routes saw the largest price decreases?",
    "Top Carriers by Passengers": "Which carriers carry the most passengers?",
    "Bottom Carriers by Passengers": "Which carriers carry the fewest passengers?",
    "Top Carriers by Price Efficiency": "Which carriers are priced highest per mile?",
    "Top Carriers by Price-Per-Mile Growth": "Which carriers are growing price per mile fastest?",
    "Top Carriers by Market Share": "Which carriers hold the largest share of their routes?",
}

# Overall Consensus's four subsection labels, in display order - shared
# between build_pdf() (rendering) and ai_layer/summarize.py's
# _build_consensus_template()/_build_consensus_claude() (generation), which
# both return a dict with these same keys plus "opening". These three are
# rendered as an Insight/Takeaway pair each (a dict of {insight, takeaway});
# "Quarterly Trends" is handled separately below since it stays a single
# plain-string paragraph, not split into Insight/Takeaway.
CONSENSUS_INSIGHT_SUBSECTIONS = ["Carrier Performance", "Price Trends", "Route Performance"]


def latest_quarter(conn) -> tuple:
    with conn.cursor() as cur:
        cur.execute("SELECT MAX(year * 10 + quarter) FROM route_summary;")
        (key,) = cur.fetchone()
    return key // 10, key % 10


def fetch_base_metrics(conn, year: int, quarter: int) -> dict:
    """Run exactly one query per entity - one for every route, one for
    every carrier - each already carrying every metric column on every
    row. Both the xlsx (the full list as-is) and all of the pdf's
    per-metric best/worst sections (sorted + head()/tail() slices, plus
    add_price_multiplier() for the two Efficiency sections - see
    _pdf_sections) are derived from these results - no metric is ever
    queried on its own. Does not include the quarterly trend, which is a
    separate whole-dataset query (see get_quarterly_trend)."""
    return {
        "Route Summary": summarize.get_top_routes(conn, year, quarter, top_n=None),
        "Carrier Summary": summarize.get_ranked_carriers(
            conn, year, quarter, "total_passengers", ascending=False, top_n=None
        ),
    }


def _pdf_sections(base_metrics: dict, carrier_growth_df: pd.DataFrame, top_routes_n: int) -> dict:
    """Derive the PDF's domain-grouped sections from the Route Summary /
    Carrier Summary dataframes by sorting on the relevant metric in
    pandas, then head()/tail() slicing - no additional queries beyond the
    already-fetched carrier_growth_df (carrier-level price-per-mile QoQ
    growth has no equivalent column on either base dataframe, so it's
    computed once in run_reports() via
    summarize.get_ranked_carrier_price_per_mile_growth() and passed in).
    Growth %/price change % sorts drop rows with no value for that column
    first (thin/new routes with no prior-quarter comparison), matching
    what the narrative engines exclude. "Worst" sections use a reversed
    tail() so the single worst performer leads, mirroring how "best"
    sections lead with the single best performer."""
    route_n = summarize.BEST_WORST_TOP_N
    efficiency_n = summarize.ROUTE_EFFICIENCY_TOP_N
    carrier_n = summarize.CARRIER_BEST_WORST_TOP_N
    carrier_share_n = summarize.CARRIER_SHARE_TOP_N

    def best_worst(df, metric, n):
        ranked = df.dropna(subset=[metric]).sort_values(metric, ascending=False)
        return ranked.head(n), ranked.tail(n).iloc[::-1]

    routes = base_metrics["Route Summary"]  # already sorted by passengers descending
    carriers = base_metrics["Carrier Summary"]  # already sorted by total_passengers descending
    # Same minimum-passenger floors as summarize.py's gather_facts(). For
    # carriers: keeps a carrier with a single low-volume appearance (which
    # can average out to a misleading 100% market share) from ranking as a
    # top performer by price efficiency or average market share. For
    # routes: keeps a route that's still small in absolute terms from
    # ranking as a "top"/"bottom" performer by growth %, price change %, or
    # price efficiency, even if it cleared transform.py's separate
    # prior-quarter floor. Neither filter applies to the by-passengers/
    # by-volume rankings, where showing genuinely small routes/carriers is
    # the point.
    qualifying_routes = routes[routes["passengers"] >= summarize.MIN_ROUTE_PASSENGERS]
    qualifying_carriers = carriers[carriers["total_passengers"] >= summarize.MIN_CARRIER_PASSENGERS]

    best_growth, worst_growth = best_worst(qualifying_routes, "passenger_growth_pct", route_n)
    best_price, worst_price = best_worst(qualifying_routes, "price_change_pct", route_n)

    # passenger_growth_pct is dropped here (not just left unformatted) -
    # the efficiency tables were overflowing the page's right margin at 10
    # columns; this is the one column that isn't part of the
    # distance-adjusted efficiency story, so it comes out rather than
    # trimming anything from the still-9-column result.
    efficiency_routes = summarize.add_price_multiplier(
        qualifying_routes.dropna(subset=["average_price", "price_per_mile"])
    ).drop(columns=["passenger_growth_pct"])
    best_efficiency, worst_efficiency = best_worst(efficiency_routes, "price_multiplier", efficiency_n)

    best_carrier_pax, worst_carrier_pax = best_worst(carriers, "total_passengers", carrier_n)
    best_carrier_efficiency = (
        qualifying_carriers.dropna(subset=["avg_price_per_mile"])
        .sort_values("avg_price_per_mile", ascending=False)
        .head(carrier_n)
    )
    best_carrier_share = (
        qualifying_carriers.dropna(subset=["avg_route_share_pct"])
        .sort_values("avg_route_share_pct", ascending=False)
        .head(carrier_share_n)
    )

    return {
        "Top Routes by Volume": routes.head(top_routes_n),
        "Best Passenger Growth": best_growth,
        "Worst Passenger Growth": worst_growth,
        "Best Price Movement": best_price,
        "Worst Price Movement": worst_price,
        "Best Price Efficiency": best_efficiency,
        "Worst Price Efficiency": worst_efficiency,
        "Top Carriers by Passengers": best_carrier_pax,
        "Bottom Carriers by Passengers": worst_carrier_pax,
        "Top Carriers by Price Efficiency": best_carrier_efficiency,
        "Top Carriers by Price-Per-Mile Growth": carrier_growth_df,
        "Top Carriers by Market Share": best_carrier_share,
    }


# ---------------------------------------------------------------------
# quarterly_report.xlsx
# ---------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _write_sheet(wb: Workbook, title: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title=title[:31])  # Excel sheet names are capped at 31 chars
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for row in df.itertuples(index=False):
        ws.append(list(row))

    for i, col in enumerate(df.columns, start=1):
        width = max(12, len(str(col)) + 2)
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"


def build_xlsx(quarterly_trend_df: pd.DataFrame, base_metrics: dict) -> None:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    _write_sheet(wb, "Quarterly Trend", quarterly_trend_df)
    for title, df in base_metrics.items():
        _write_sheet(wb, title, df)

    wb.save(XLSX_FILE)
    sheets = [("Quarterly Trend", quarterly_trend_df), *base_metrics.items()]
    row_counts = ", ".join(f"{title}={len(df):,}" for title, df in sheets)
    print(f"Wrote {XLSX_FILE} ({len(sheets)} sheets, no row limit)\n  {row_counts}")


# ---------------------------------------------------------------------
# executive_summary.pdf
# ---------------------------------------------------------------------

def _escape_for_pdf(text: str) -> str:
    """reportlab's Paragraph parses its text as a small XML-like markup
    language, so an unescaped &, <, or > in AI-written blurb text (unlikely,
    but not impossible) would break rendering. Escape before wrapping any
    narrative text in a Paragraph."""
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _format_float_cell(v, is_dollar: bool, is_percent: bool, is_multiplier: bool = False) -> str:
    """One float-column cell as a display string - $ prefix for dollar
    columns (DOLLAR_COLS), % suffix for percent columns (PERCENT_COLS,
    already stored as e.g. 5.5 meaning 5.5%), "x" suffix for multiplier
    columns (MULTIPLIER_COLS, a ratio vs. distance-band peers, e.g.
    "2.75x"), plain comma-formatted number otherwise. NaN (no prior-quarter
    comparison, etc.) renders as an empty cell rather than "nan"."""
    if pd.isna(v):
        return ""
    if is_dollar:
        return f"${v:,.2f}"
    if is_percent:
        return f"{v:,.2f}%"
    if is_multiplier:
        return f"{v:,.2f}x"
    return f"{v:,.2f}"


def _df_to_table(df: pd.DataFrame, float_cols=()) -> Table:
    display = df.copy()
    for col in float_cols:
        if col in display.columns:
            is_dollar = col in DOLLAR_COLS
            is_percent = col in PERCENT_COLS
            is_multiplier = col in MULTIPLIER_COLS
            display[col] = display[col].map(
                lambda v: _format_float_cell(v, is_dollar, is_percent, is_multiplier)
            )
    data = [list(display.columns)] + display.astype(str).values.tolist()

    table = Table(data, repeatRows=1)
    table.hAlign = "LEFT"  # reportlab centers tables by default - keep them left-aligned instead
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ]
        )
    )
    # Never split a table mid-row across a page break - push the whole thing
    # onto the next page instead.
    table.splitByRow = 0
    return table


def build_pdf(narrative: dict, quarterly_trend_df: pd.DataFrame, sections: dict) -> None:
    doc = SimpleDocTemplate(str(PDF_FILE), pagesize=letter)
    styles = getSampleStyleSheet()
    # Section-level headers (Carrier Performance, Price Trends, Route
    # Performance, and Overall Consensus) - smaller than the 18pt report
    # Title, bigger than the 10pt table sub-headings below, left-aligned,
    # and with spaceBefore/spaceAfter zeroed out - by default a Paragraph
    # built on Heading1 carries its own built-in spacing on top of whatever
    # Spacer we place around it, which was silently stacking with our
    # explicit spacers and blowing out the gaps; from here on, spacing is
    # only ever controlled by our own explicit Spacer() calls.
    section_heading_style = ParagraphStyle(
        "SectionHeading", parent=styles["Heading1"], fontSize=15, leading=18,
        spaceBefore=0, spaceAfter=0,
    )
    # Smaller style for per-table titles ("Top Routes by Volume", etc.) and
    # the Overall Consensus subsection labels ("Carrier Performance",
    # etc.) - same spaceBefore/spaceAfter zeroing as above, for the same
    # reason.
    table_title_style = ParagraphStyle(
        "TableTitle", parent=styles["Heading2"], fontSize=10, leading=12,
        spaceBefore=0, spaceAfter=2,
    )
    # One-line business question under each sub-header (see
    # SECTION_QUESTIONS) - italic, smaller than the 10pt sub-header so it
    # reads as a caption, not another heading level.
    question_style = ParagraphStyle(
        "TableQuestion", parent=styles["Italic"], fontSize=8, leading=10, spaceAfter=4,
    )
    # Style for the AI-written blurb placed under each table - smaller than
    # normal body text so it visually reads as commentary on the table above
    # it, not a new independent section.
    blurb_style = ParagraphStyle(
        "Blurb", parent=styles["Normal"], fontSize=9, leading=12, spaceBefore=4,
    )
    # "Insight:"/"Takeaway:" labels within the Overall Consensus - inline
    # italic via <i> (reportlab's Paragraph supports a small XML-like
    # markup subset), not bold, so they read as a distinct lighter tier
    # under the bold subsection headers above them instead of blending
    # into the same visual weight.
    consensus_body_style = ParagraphStyle(
        "ConsensusBody", parent=styles["Normal"], fontSize=10, leading=13,
    )

    def blurb_paragraph(title: str) -> Paragraph:
        return Paragraph(_escape_for_pdf(narrative.get(title, "")), blurb_style)

    # Latest quarter tracked in the report - same value used throughout
    # (quarterly_trend_df is ordered by year, quarter ascending, so the
    # last row is the current one) - stamped into the page header on every
    # page via _draw_page_header() below, instead of repeating it next to
    # each section heading.
    latest_row = quarterly_trend_df.iloc[-1]
    latest_year, latest_quarter = int(latest_row["year"]), int(latest_row["quarter"])

    def _draw_page_header(canvas, doc) -> None:
        """Runs once per page (both the first page and every later page,
        via doc.build()'s onFirstPage/onLaterPages below) - draws the
        latest quarter into the top margin, right-aligned, so it's visible
        on every page like a running header rather than repeated inline
        next to each section heading."""
        canvas.saveState()
        canvas.setFont("Helvetica-Oblique", 9)
        canvas.setFillColor(colors.grey)
        canvas.drawRightString(
            doc.pagesize[0] - doc.rightMargin, doc.pagesize[1] - 0.5 * inch, f"{latest_year} Q{latest_quarter}"
        )
        canvas.restoreState()

    story = [
        Paragraph(REPORT_TITLE, styles["Title"]),
        Spacer(1, 0.15 * inch),
        Paragraph(REPORT_OVERVIEW, styles["Normal"]),
        Spacer(1, 0.25 * inch),
    ]

    # Quarterly Trends: subheader, business question, table, a real gap,
    # then its blurb - kept together so nothing can get separated from its
    # table across a page break.
    story.append(
        KeepTogether(
            [
                Paragraph("Quarterly Trends", table_title_style),
                Paragraph(SECTION_QUESTIONS["Quarterly Trends"], question_style),
                Spacer(1, 0.1 * inch),
                _df_to_table(
                    quarterly_trend_df.tail(QUARTERLY_TREND_PDF_LIMIT), float_cols=QUARTERLY_TREND_FLOAT_COLS
                ),
                Spacer(1, 0.15 * inch),
                blurb_paragraph("Quarterly Trends"),
            ]
        )
    )

    # Carrier Performance / Price Trends / Route Performance each always
    # start on a fresh page - a forced PageBreak before the domain heading,
    # not just a Spacer, so a domain never starts mid-page the way PRICE did
    # before this fix. Each table gets subheader, table, a real gap, then
    # its matching blurb right below it, all kept together so a blurb can
    # never get separated from its table across a page break.
    for domain, section_titles in DOMAIN_GROUPS:
        for i, title in enumerate(section_titles):
            float_cols = SECTION_FLOAT_COLS[title]
            block = []
            if i == 0:
                story.append(PageBreak())
                block.append(Paragraph(domain, section_heading_style))
            block.append(Spacer(1, 0.15 * inch))
            block.append(Paragraph(title, table_title_style))
            block.append(Paragraph(SECTION_QUESTIONS[title], question_style))
            block.append(Spacer(1, 0.1 * inch))
            block.append(_df_to_table(sections[title], float_cols=float_cols))
            block.append(Spacer(1, 0.15 * inch))
            block.append(blurb_paragraph(title))
            story.append(KeepTogether(block))

    # Overall Consensus - always starts on its own page. A short opening
    # overview, an Insight/Takeaway pair each for Carrier Performance, Price
    # Trends, and Route Performance, then a single Quarterly Trends
    # paragraph (see ai_layer/summarize.py's
    # _build_consensus_template()/_build_consensus_claude()). Spacing here
    # is intentionally tight - shorter gaps than the rest of the report -
    # so the whole section has the best chance of fitting on one page
    # given the Claude engine's 3-4 sentence cap on each piece.
    story.append(PageBreak())
    story.append(Paragraph("Overall Consensus", section_heading_style))
    story.append(Spacer(1, 0.1 * inch))

    consensus = narrative.get("Overall Consensus", {})
    if not isinstance(consensus, dict):
        consensus = {"opening": consensus}  # defensive fallback for a plain-string blurb

    opening_text = consensus.get("opening", "")
    if opening_text:
        story.append(Paragraph(_escape_for_pdf(opening_text), styles["Normal"]))
        story.append(Spacer(1, 0.15 * inch))

    for label in CONSENSUS_INSIGHT_SUBSECTIONS:
        sub = consensus.get(label, {})
        if not isinstance(sub, dict):
            sub = {"insight": sub, "takeaway": ""}  # defensive fallback for a plain-string blurb
        insight_text = sub.get("insight", "")
        takeaway_text = sub.get("takeaway", "")
        if not insight_text and not takeaway_text:
            continue
        block = [Paragraph(label, table_title_style), Spacer(1, 0.05 * inch)]
        if insight_text:
            block.append(Paragraph(f"<i>Insight:</i> {_escape_for_pdf(insight_text)}", consensus_body_style))
        if takeaway_text:
            block.append(Spacer(1, 0.04 * inch))
            block.append(Paragraph(f"<i>Takeaway:</i> {_escape_for_pdf(takeaway_text)}", consensus_body_style))
        block.append(Spacer(1, 0.1 * inch))
        story.append(KeepTogether(block))

    quarterly_text = consensus.get("Quarterly Trends", "")
    if quarterly_text:
        story.append(
            KeepTogether(
                [
                    Paragraph("Quarterly Trends", table_title_style),
                    Spacer(1, 0.05 * inch),
                    Paragraph(_escape_for_pdf(quarterly_text), consensus_body_style),
                ]
            )
        )

    doc.build(story, onFirstPage=_draw_page_header, onLaterPages=_draw_page_header)
    print(f"Wrote {PDF_FILE} (quarterly trend + {len(sections)} tables across {len(DOMAIN_GROUPS)} domains, plus Overall Consensus)")


def run_reports(top_n: int = 10) -> None:
    """Build quarterly_report.xlsx (3 sheets, every row, no limit) and
    executive_summary.pdf (quarterly trend + capped, domain-grouped
    sections) from whatever is currently in route_summary /
    carrier_summary / price_variance_summary. Callable directly (e.g. from
    pipeline/main.py) as well as via the CLI."""
    conn = get_connection()
    try:
        year, quarter = latest_quarter(conn)
        print(f"Building reports for {year} Q{quarter}...")

        quarterly_trend_df = summarize.get_quarterly_trend(conn)
        base_metrics = fetch_base_metrics(conn, year, quarter)
        carrier_growth_df = summarize.get_ranked_carrier_price_per_mile_growth(
            conn, year, quarter, ascending=False, top_n=summarize.CARRIER_BEST_WORST_TOP_N,
            min_passengers=summarize.MIN_CARRIER_PASSENGERS,
        )
        pdf_sections = _pdf_sections(base_metrics, carrier_growth_df, top_routes_n=top_n)

        narrative = summarize.generate_narrative(conn, top_n=top_n)
    finally:
        conn.close()

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    build_xlsx(quarterly_trend_df, base_metrics)
    build_pdf(narrative, quarterly_trend_df, pdf_sections)


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate quarterly_report.xlsx and executive_summary.pdf")
    parser.add_argument(
        "--top",
        type=int,
        default=10,
        help="Rows in the PDF's 'Top Routes by Volume' table (default: 10). "
        "The xlsx has no row limit on any sheet; every other ROUTE/PRICE/EFFICIENCY table is capped at "
        "ai_layer/summarize.py's BEST_WORST_TOP_N (currently 3), and every CARRIER table at "
        "CARRIER_BEST_WORST_TOP_N (currently 5).",
    )
    args = parser.parse_args()

    run_reports(top_n=args.top)


if __name__ == "__main__":
    main()
